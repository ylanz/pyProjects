import sys
from time import sleep

#MODULOS A INSTALAR
try:
    import openpyxl as lexcel
except ImportError:
    print("Modulo 'openpyxl' no instalado")
    sys.exit()
    
try:    
    import pymysql as cnn
except ImportError:
    print("Modulo 'pymysql' no instalado")
    sys.exit()
    
try: 
    import progressbar
except ImportError:
    print("Modulo 'progressbar' no instalado")
    sys.exit()
#***********-------------

#Variables globales
global db
global pProductos
global Categorias
global pSubCategorias
global pUND


def cargarLibroExcel(libroName="datosv2.xlsx"):
    #cargar el libro
    print("Cargando libro...")
    fexl = lexcel.load_workbook(libroName, data_only=True)
    print("Cargando pestañas...")
    #**pestañas del libro
    #codigo del producto, descripcion, codigo UND, descripcion UND, codigo categoria, descripcion categoria
    #cod de subcategoria, descripcion subcategoria    
    global pProductos
    global Categorias
    global pSubCategorias
    global pUND
    pProductos = fexl["Productos"]
    #Descripcion, codigo
    #pSubCategorias = fexl["SUB_CATEGORIA"]
    pSubCategorias = fexl["relacion cat-subcat"]
    #Descripcion, codigo
    pUND = fexl["Unidades de Medida"]
    #Descripcion, codigo
    Categorias = fexl["CATEGORIAS"]
 
    #Comprobando carga de Unidades
    #print("Datos cargados de las pestañas...")
    fila = col = 0
    a = ""
    #for fila in range (2, pUND.max_row):
    #    for col in range(1, 3):
    #      a+= "\t" + str(pUND.cell(fila,col).value) + "\t"
    #    print(a)
    #    a = ""
    #print("")
    for fila in range (2, pSubCategorias.max_row):
        for col in range(4, 6):
          a+= str(pSubCategorias.cell(fila,col).value) + "\t"
        print(a)
        a = ""  
    #for fila in range (2, Categorias.max_row):
    #    for col in range(1, 3):
    #      a+= "\t" + str(Categorias.cell(fila,col).value) + "\t"
    #    print(a)
    #    a = "" 
    fila = col = 0
    a = ""
    bar = progressbar.ProgressBar(maxval = pProductos.max_row, \
    widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
    bar.start()
    
    for fila in range (2, pProductos.max_row):
        for col in range(1, 3):
            a = col
        bar.update(fila + 1)
        
    bar.finish()
        
        
def conectarBD():
    #Conectando a la BD
    print("")
    print("Conectando a la BD...")
    global db
    db = cnn.connect(host = '10.1.1.32',
                     user = 'root',
                     password = '4c3r04dm1n',
                     db = 'intranet',
                     charset = 'utf8mb4',
                     cursorclass = cnn.cursors.DictCursor)
    print("***Conexion Exitosa")

def insertarDatosUnidadMedida(nameTabla="adm_unidad_medidas_copy"):
    print("")
    print("Insertando en la BD...")
    global db
    global pUND
    
    try:
        with db.cursor() as cursor:
            bar = progressbar.ProgressBar(maxval=pUND.max_row, \
            widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
            bar.start()
            
            for fila in range (2, pUND.max_row):
                for col in range(1, 3):
                    abc= 0
                #sql = f"INSERT INTO {nameTabla} (nombre, abrev, orden, idAdmTipoMedida) values (%s, %s, %s, %s)"
                #cursor.execute(sql, (pUND.cell(fila, 1).value, "", fila, 2))
                bar.update(fila + 1)
                #sleep(0.1)
            
            bar.finish()
            db.commit()
            print("***Filas insertadas: "+ str(pUND.max_row))
            print(f"***Insercion en la tabla {nameTabla} existoso!")
        #with db.cursor() as cursor: 
        #    cursor.execute("SELECT * FROM adm_unidad_medidas_copy")
        #    print(cursor.fetchone())
    finally:
        db.close()

def insertarDatosCategorias(nameTabla="adm_grupos_productos_copy"):
    print("")
    print("Insertando en la BD...")
    global db
    global Categorias
    
    try:
        with db.cursor() as cursor:
            bar = progressbar.ProgressBar(maxval=Categorias.max_row, \
            widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
            bar.start()
            
            for fila in range (2, Categorias.max_row):
                for col in range(1, 3):
                    abc= 0
                sql = f"INSERT INTO {nameTabla} (nombre, orden) values (%s, %s)"
                cursor.execute(sql, (Categorias.cell(fila, 2).value, fila))
                bar.update(fila + 1)
                #sleep(0.1)
            
            bar.finish()
            db.commit()
            print("***Filas insertadas: "+ str(Categorias.max_row))
            print(f"***Insercion en la tabla {nameTabla} existoso!")
        #with db.cursor() as cursor: 
        #    cursor.execute("SELECT * FROM adm_unidad_medidas_copy")
        #    print(cursor.fetchone())
    finally:
        db.close()

def insertarDatosSubCategorias(nameTabla="adm_sub_grupos_productos_copy"):
    print("")
    print("Insertando en la BD...")
    global db
    global pSubCategorias
    
    try:
        with db.cursor() as cursor:
            bar = progressbar.ProgressBar(maxval=pSubCategorias.max_row, \
            widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
            bar.start()
            
            for fila in range (2, pSubCategorias.max_row):
                #for col in range(4, 6):
                #    abc= 0
                sql = f"INSERT INTO {nameTabla} (nombre, orden, idAdmGrupoProducto) values (%s, %s, %s)"
                cursor.execute(sql, (pSubCategorias.cell(fila, 5).value, fila, pSubCategorias.cell(fila, 4).value))
                bar.update(fila + 1)
                #sleep(0.1)
            
            bar.finish()
            db.commit()
            print("***Filas insertadas: "+ str(pSubCategorias.max_row))
            print(f"***Insercion en la tabla {nameTabla} existoso!")
        #with db.cursor() as cursor: 
        #    cursor.execute("SELECT * FROM adm_unidad_medidas_copy")
        #    print(cursor.fetchone())
    finally:
        db.close()


def insertarDatosProducto(nameTabla="adm_productos_copy"):
    print("")
    print("Insertando en la BD...")
    global db
    global pProductos
    
    try:
        with db.cursor() as cursor:
            bar = progressbar.ProgressBar(maxval=pProductos.max_row, \
            widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
            bar.start()
            
            for fila in range (2, pProductos.max_row):
                for col in range(1, 3):
                    abc= 0
                sql = f"INSERT INTO {nameTabla} (codigo, nombre, idAdmUnidadMedida, idAdmGrupoProducto, idAdmSubGrupoProducto) values (%s, %s, %s, %s, %s)"
                cursor.execute(sql, (pProductos.cell(fila, 1).value,pProductos.cell(fila, 2).value, 
                                    pProductos.cell(fila, 3).value, pProductos.cell(fila, 5).value,
                                    pProductos.cell(fila, 7).value))
                bar.update(fila + 1)
                #sleep(0.1)
            
            bar.finish()
            db.commit()
            print("***Filas insertadas: "+ str(pProductos.max_row))
            print(f"***Insercion en la tabla {nameTabla} existoso!")
        #with db.cursor() as cursor: 
        #    cursor.execute("SELECT * FROM adm_unidad_medidas_copy")
        #    print(cursor.fetchone())
    finally:
        db.close()


cargarLibroExcel()  
conectarBD()

#insertarDatosUnidadMedida()            
#insertarDatosCategorias()
insertarDatosSubCategorias()
#insertarDatosProducto()
